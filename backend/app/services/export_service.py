"""
app/services/export_service.py
────────────────────────────────
Generates downloadable reports from stored KPI data.

Structure mirrors the original JS exportManager.js:
  • xlsx — 4 sheets: Summary KPIs / Daily Sales / Batch Performance / Category Performance
  • pdf  — A4, reportlab, readable tables (NOT a raw JSON dump)

Both formats accept the kpi_json dict returned by calculate_kpis() and an optional
dataset metadata dict for headers/footers.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, Optional

# ── xlsx ─────────────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

# ── pdf ──────────────────────────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


# ═══════════════════════════════════════════════════════════════════════════════
# xlsx export
# ═══════════════════════════════════════════════════════════════════════════════

_HEADER_FILL = PatternFill("solid", fgColor="0F172A")
_SUBHEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
_ALT_FILL = PatternFill("solid", fgColor="F1F5F9")
_HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
_BODY_FONT = Font(name="Calibri", size=10)
_BOLD_FONT = Font(name="Calibri", size=10, bold=True)
_THIN = Side(border_style="thin", color="CBD5E1")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _write_header_row(ws, row: int, values: list[str]) -> None:
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_data_row(ws, row: int, values: list, alt: bool = False) -> None:
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = _BODY_FONT
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="right" if isinstance(val, (int, float)) else "left")
        if alt:
            cell.fill = _ALT_FILL


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def generate_xlsx(kpi_json: Dict[str, Any], filename: str = "dataset") -> bytes:
    """
    Build an Excel workbook from stored KPI data.
    Sheets: Summary KPIs | Daily Sales | Batch Performance | Category Performance
    Returns raw bytes suitable for a StreamingResponse.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    # ── 1. Summary KPIs ──────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Summary KPIs")
    kpi_rows = [
        ("Total Gross Sales",         kpi_json.get("gross_revenue")),
        ("Total Net Revenue",         kpi_json.get("net_revenue")),
        ("Refund Amount",             kpi_json.get("refund_amount")),
        ("Refund %",                  kpi_json.get("refund_percent")),
        ("Total Orders",              kpi_json.get("orders")),
        ("Average Selling Price",     kpi_json.get("average_selling_price")),
        ("Daily Run Rate (DRR)",      kpi_json.get("daily_run_rate")),
        ("Today's Sales",             kpi_json.get("today_sales")),
        ("Yesterday's Sales",         kpi_json.get("yesterday_sales")),
        ("WoW Growth %",              kpi_json.get("wow_growth")),
        ("MoM Growth %",              kpi_json.get("mom_growth")),
        ("Days Compared (MoM)",       kpi_json.get("days_compared")),
        ("Total Target",              kpi_json.get("total_target")),
        ("Target Achieved %",         kpi_json.get("target_achieved_percent")),
        ("Target Remaining",          kpi_json.get("target_remaining")),
        ("Required DRR to hit target",kpi_json.get("required_drr")),
        ("Forecast Month Revenue",    kpi_json.get("expected_month_revenue")),
        ("Forecast Target Achievement %", kpi_json.get("forecast_target_achievement")),
        ("Total Profit",              kpi_json.get("profit")),
        ("Total Loss (refunds)",      kpi_json.get("loss")),
    ]
    _write_header_row(ws1, 1, ["Metric", "Value"])
    for i, (metric, value) in enumerate(kpi_rows, start=2):
        _write_data_row(ws1, i, [metric, value], alt=(i % 2 == 0))
    _auto_width(ws1)

    # ── 2. Category Performance ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Category Performance")
    cat_perf = kpi_json.get("category_performance", [])
    _write_header_row(ws2, 1, ["Category", "Revenue", "Orders", "Profit", "Margin Used", "Is Default Margin"])
    for i, row in enumerate(cat_perf, start=2):
        _write_data_row(ws2, i, [
            row.get("category"),
            row.get("revenue"),
            row.get("orders"),
            row.get("profit"),
            row.get("margin_used"),
            row.get("is_default"),
        ], alt=(i % 2 == 0))
    _auto_width(ws2)

    # ── 3. Batch Performance ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("Batch Performance")
    batch_perf = kpi_json.get("batch_performance", [])
    _write_header_row(ws3, 1, ["Batch", "Revenue", "Admissions", "Capacity", "Fill %", "Profit", "Refund Amount", "Capacity Missing"])
    for i, row in enumerate(batch_perf, start=2):
        _write_data_row(ws3, i, [
            row.get("batch"),
            row.get("revenue"),
            row.get("admissions"),
            row.get("capacity"),
            row.get("fill_percent"),
            row.get("profit"),
            row.get("refund_amount"),
            row.get("capacity_missing"),
        ], alt=(i % 2 == 0))
    _auto_width(ws3)

    # ── 4. Teacher & State Performance ───────────────────────────────────────
    ws4 = wb.create_sheet("Teacher & State")
    teacher_perf = kpi_json.get("teacher_performance", [])
    _write_header_row(ws4, 1, ["Teacher", "Revenue", "Admissions"])
    for i, row in enumerate(teacher_perf, start=2):
        _write_data_row(ws4, i, [row.get("teacher"), row.get("revenue"), row.get("admissions")], alt=(i % 2 == 0))

    state_start = len(teacher_perf) + 4
    state_perf = kpi_json.get("state_performance", [])
    _write_header_row(ws4, state_start, ["State", "Revenue"])
    for i, row in enumerate(state_perf, start=state_start + 1):
        _write_data_row(ws4, i, [row.get("state"), row.get("revenue")], alt=(i % 2 == 0))
    _auto_width(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# PDF export
# ═══════════════════════════════════════════════════════════════════════════════

_NAVY = colors.HexColor("#0F172A")
_BLUE = colors.HexColor("#1E3A5F")
_LIGHT = colors.HexColor("#F1F5F9")
_WHITE = colors.white
_ACCENT = colors.HexColor("#3B82F6")


def _pdf_table_style(has_header: bool = True) -> TableStyle:
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), _BLUE),
        ("TEXTCOLOR",  (0, 0), (-1, 0), _WHITE),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 9),
        ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_WHITE, _LIGHT]),
        ("FONTNAME",   (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",   (0, 1), (-1, -1), 8),
        ("GRID",       (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
    ]
    return TableStyle(style)


def _fmt(val, prefix="₹", suffix="", pct=False):
    if val is None:
        return "N/A"
    if pct:
        return f"{val:.2f}%"
    if isinstance(val, float):
        return f"{prefix}{val:,.2f}{suffix}"
    return str(val)


def generate_pdf(kpi_json: Dict[str, Any], filename: str = "dataset", dataset_meta: Optional[Dict] = None) -> bytes:
    """
    Build an A4 PDF report from stored KPI data using reportlab.
    Sections: header / summary KPIs / category performance / batch performance / teacher performance
    Returns raw bytes suitable for a StreamingResponse.
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title", parent=styles["Title"],
        textColor=_NAVY, fontSize=18, spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "subtitle", parent=styles["Normal"],
        textColor=colors.HexColor("#475569"), fontSize=9, spaceAfter=16,
    )
    section_style = ParagraphStyle(
        "section", parent=styles["Heading2"],
        textColor=_NAVY, fontSize=12, spaceBefore=14, spaceAfter=6,
        borderPad=2,
    )

    story = []

    # ── Header ────────────────────────────────────────────────────────────────
    story.append(Paragraph("BI Dashboard — Executive Report", title_style))
    generated_at = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    net_rev = _fmt(kpi_json.get("net_revenue"))
    uploaded_file = (dataset_meta or {}).get("filename", filename)
    story.append(Paragraph(
        f"Generated: {generated_at} &nbsp;|&nbsp; Source file: {uploaded_file} &nbsp;|&nbsp; Net Revenue: {net_rev}",
        subtitle_style,
    ))

    # ── 1. Summary KPIs ───────────────────────────────────────────────────────
    story.append(Paragraph("Summary KPIs", section_style))
    summary_data = [
        ["Metric", "Value"],
        ["Gross Revenue",              _fmt(kpi_json.get("gross_revenue"))],
        ["Net Revenue",                _fmt(kpi_json.get("net_revenue"))],
        ["Refund Amount",              _fmt(kpi_json.get("refund_amount"))],
        ["Refund %",                   _fmt(kpi_json.get("refund_percent"), prefix="", suffix="%")],
        ["Total Orders",               str(kpi_json.get("orders", 0))],
        ["Average Selling Price",      _fmt(kpi_json.get("average_selling_price"))],
        ["Daily Run Rate",             _fmt(kpi_json.get("daily_run_rate"))],
        ["Today's Sales",              _fmt(kpi_json.get("today_sales"))],
        ["Yesterday's Sales",          _fmt(kpi_json.get("yesterday_sales"))],
        ["WoW Growth",                 _fmt(kpi_json.get("wow_growth"), prefix="", suffix="%")],
        ["MoM Growth",                 _fmt(kpi_json.get("mom_growth"), prefix="", suffix="%")],
        ["Total Target",               _fmt(kpi_json.get("total_target"))],
        ["Target Achieved",            _fmt(kpi_json.get("target_achieved_percent"), prefix="", suffix="%")],
        ["Target Remaining",           _fmt(kpi_json.get("target_remaining"))],
        ["Required DRR",               _fmt(kpi_json.get("required_drr"))],
        ["Total Profit",               _fmt(kpi_json.get("profit"))],
        ["Total Loss (Refunds)",       _fmt(kpi_json.get("loss"))],
    ]
    col_w = [(doc.width - 10) * 0.6, (doc.width - 10) * 0.4]
    t = Table(summary_data, colWidths=col_w)
    t.setStyle(_pdf_table_style())
    story.append(t)

    # ── 2. Category Performance ───────────────────────────────────────────────
    cat_perf = kpi_json.get("category_performance", [])
    if cat_perf:
        story.append(Paragraph("Category Performance", section_style))
        cat_data = [["Category", "Revenue", "Orders", "Profit", "Margin", "Est?"]]
        for row in cat_perf:
            cat_data.append([
                row.get("category", ""),
                _fmt(row.get("revenue")),
                str(row.get("orders", 0)),
                _fmt(row.get("profit")),
                f"{row.get('margin_used', 0):.0%}",
                "Yes" if row.get("is_default") else "No",
            ])
        cat_col_w = [doc.width / 6] * 6
        tc = Table(cat_data, colWidths=cat_col_w)
        tc.setStyle(_pdf_table_style())
        story.append(tc)

    # ── 3. Batch Performance ──────────────────────────────────────────────────
    batch_perf = kpi_json.get("batch_performance", [])
    if batch_perf:
        story.append(Paragraph("Batch Performance", section_style))
        batch_data = [["Batch", "Revenue", "Admissions", "Fill %", "Profit"]]
        for row in batch_perf:
            fill = _fmt(row.get("fill_percent"), prefix="", suffix="%") if row.get("fill_percent") is not None else "N/A"
            batch_data.append([
                row.get("batch", ""),
                _fmt(row.get("revenue")),
                str(row.get("admissions", 0)),
                fill,
                _fmt(row.get("profit")),
            ])
        batch_col_w = [doc.width / 5] * 5
        tb = Table(batch_data, colWidths=batch_col_w)
        tb.setStyle(_pdf_table_style())
        story.append(tb)

    # ── 4. Teacher Performance ────────────────────────────────────────────────
    teacher_perf = kpi_json.get("teacher_performance", [])
    if teacher_perf:
        story.append(Paragraph("Teacher Performance", section_style))
        teacher_data = [["Teacher", "Revenue", "Admissions"]]
        for row in teacher_perf:
            teacher_data.append([
                row.get("teacher", ""),
                _fmt(row.get("revenue")),
                str(row.get("admissions", 0)),
            ])
        teach_col_w = [doc.width / 3] * 3
        tt = Table(teacher_data, colWidths=teach_col_w)
        tt.setStyle(_pdf_table_style())
        story.append(tt)

    doc.build(story)
    return buf.getvalue()
